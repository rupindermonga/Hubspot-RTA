import streamlit as st
import pandas as pd
import re
import io
import bcrypt
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

MAX_UPLOAD_ROWS = 50000  # prevent memory exhaustion

# ── Server-side rate limiter (shared across ALL sessions) ──
import threading

@st.cache_resource
def _get_rate_limiter():
    """Singleton rate limiter persisted across sessions via cache_resource."""
    return {
        'lock': threading.Lock(),
        'attempts': {},   # username -> list of timestamps
        'global': [],     # all failed timestamps (IP-agnostic but still global)
    }

RATE_WINDOW = 300       # 5-minute sliding window
MAX_PER_USER = 5        # max failures per username in window
MAX_GLOBAL = 30         # max total failures across all users in window

def check_rate_limit(username):
    """Return (allowed, message). Enforces per-user AND global limits."""
    rl = _get_rate_limiter()
    now = time.time()
    with rl['lock']:
        # Prune old entries
        rl['global'] = [t for t in rl['global'] if now - t < RATE_WINDOW]
        if username in rl['attempts']:
            rl['attempts'][username] = [t for t in rl['attempts'][username] if now - t < RATE_WINDOW]

        # Check global limit
        if len(rl['global']) >= MAX_GLOBAL:
            return False, "Too many login failures across all accounts. Try again later."

        # Check per-user limit
        user_attempts = rl['attempts'].get(username, [])
        if len(user_attempts) >= MAX_PER_USER:
            oldest = user_attempts[0]
            wait = int(RATE_WINDOW - (now - oldest)) + 1
            return False, f"Account '{username}' locked for {wait}s due to too many failed attempts."

        return True, ""

def record_failed_attempt(username):
    """Record a failed login for rate limiting."""
    rl = _get_rate_limiter()
    now = time.time()
    with rl['lock']:
        rl['global'].append(now)
        if username not in rl['attempts']:
            rl['attempts'][username] = []
        rl['attempts'][username].append(now)

# ── Page config ──
st.set_page_config(page_title="Address Matcher", page_icon="📍", layout="wide")

# ── Authentication ──
# Credentials are loaded from Streamlit secrets (see Streamlit docs).
# Managed separately — never committed to version control.

def verify_password(password, stored_hash):
    """Check password against bcrypt hash (salted, slow, constant-time)."""
    return bcrypt.checkpw(password.encode(), stored_hash.encode())

def login():
    """Show login form with server-side rate limiting and session timeout."""
    if 'authenticated' not in st.session_state:
        st.session_state.authenticated = False
        st.session_state.username = ''
        st.session_state.login_time = 0.0

    if st.session_state.authenticated:
        # Session timeout after 8 hours
        if st.session_state.login_time > 0 and (time.time() - st.session_state.login_time) > 28800:
            st.session_state.authenticated = False
            st.session_state.username = ''
            st.warning("Session expired. Please log in again.")
        else:
            return True

    users = st.secrets.get("users", {})

    st.title("🔐 Address Matcher — Login")
    st.markdown("Please log in to continue.")

    with st.form("login_form"):
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Log in", use_container_width=True)

        if submitted:
            # Server-side rate limit check (shared across ALL sessions)
            allowed, msg = check_rate_limit(username)
            if not allowed:
                st.error(msg)
            elif username in users and verify_password(password, users[username]):
                st.session_state.authenticated = True
                st.session_state.username = username
                st.session_state.login_time = time.time()
                st.rerun()
            else:
                record_failed_attempt(username)
                st.error("Invalid username or password.")
    return False

if not login():
    st.stop()

# ── Logged in: show user + logout ──
st.sidebar.markdown(f"**Logged in as:** {st.session_state.username}")
if st.sidebar.button("Logout"):
    st.session_state.authenticated = False
    st.session_state.username = ''
    st.rerun()

st.title("📍 Address Matcher")
st.markdown("Upload **Hubspot** and **RTA** files separately. The app matches addresses and appends RTA Address + RTA Status to the Hubspot file.")

# ── SEC-06: Formula injection sanitization ──
FORMULA_PREFIXES = ('=', '+', '-', '@', '\t', '\r')

def sanitize_cell(val):
    """Prevent Excel formula injection by escaping dangerous prefixes."""
    if isinstance(val, str) and val and val[0] in FORMULA_PREFIXES:
        return "'" + val
    return val

def sanitize_dataframe(df):
    """Apply formula sanitization to all string columns in a DataFrame."""
    df_clean = df.copy()
    for col in df_clean.select_dtypes(include='object').columns:
        df_clean[col] = df_clean[col].apply(lambda v: sanitize_cell(v) if isinstance(v, str) else v)
    return df_clean

# ── Normalization engine ──
ABBREVS = {
    'STREET': 'ST', 'ROAD': 'RD', 'DRIVE': 'DR', 'AVENUE': 'AVE',
    'BOULEVARD': 'BLVD', 'CRESCENT': 'CRES', 'COURT': 'CRT', 'PLACE': 'PL',
    'LANE': 'LN', 'CIRCLE': 'CIR', 'TERRACE': 'TERR', 'HIGHWAY': 'HWY',
    'TRAIL': 'TRL', 'SQUARE': 'SQ', 'PARKWAY': 'PKY', 'WAY': 'WAY',
    'CLOSE': 'CL', 'GROVE': 'GRV', 'HEIGHTS': 'HTS', 'RIDGE': 'RDG',
    'NORTH': 'N', 'SOUTH': 'S', 'EAST': 'E', 'WEST': 'W',
    'REG': 'REGIONAL',
}

def clean_address(s):
    """Strip PO Box, RR, Suite, Apt, Unit noise."""
    s = str(s).strip()
    s = re.sub(r'^(?:P[\s.]?O[\s.]?\s*)?BOX\s*#?\s*\d*[\s,/]*', '', s, flags=re.IGNORECASE)
    s = re.sub(r'^SUITE\s+\d+\s*', '', s, flags=re.IGNORECASE)
    s = re.sub(r'^RR\s*#?\s*\d+[\s,]*', '', s, flags=re.IGNORECASE)
    s = re.sub(r'[\s,]+(?:P[\s.]?O[\s.]?\s*)?BOX\s*#?\s*\d*.*$', '', s, flags=re.IGNORECASE)
    s = re.sub(r'[\s,]+RR\s*#?\s*\d+.*$', '', s, flags=re.IGNORECASE)
    s = re.sub(r'[\s,]+(?:SUITE|APT|UNIT)\s*#?\s*\w*.*$', '', s, flags=re.IGNORECASE)
    return s.strip(' ,/')

def normalize(s):
    s = clean_address(s)
    s = s.upper().replace('.', '')
    s = re.sub(r'[^A-Z0-9 ]', '', s)
    s = re.sub(r' +', ' ', s).strip()
    words = s.split()
    words = [ABBREVS.get(w, w) for w in words]
    return ' '.join(words)

def strip_direction(s):
    # QA-02: Only strip trailing direction if street name has 3+ words
    # (prevents "123 N ST" from losing the "N" which IS the street name)
    words = s.split()
    if len(words) >= 3 and words[-1] in ('N', 'S', 'E', 'W'):
        return ' '.join(words[:-1])
    return s

def strip_unit(s):
    """QA-03: Strip unit suffixes like -U1, -U2, U1 from address numbers.
    '97U2 PIONEER RD' -> '97 PIONEER RD', '9632U1 HWY 638' -> '9632 HWY 638'"""
    return re.sub(r'^(\d+)[- ]?U\d+', r'\1', s)

def norm_pc(s):
    s = str(s).strip().upper().replace(' ', '')
    # QA-01: Treat NaN/empty/N/A as empty — prevents false matches
    if s in ('NAN', 'NONE', 'N/A', 'NA', 'NULL', ''):
        return ''
    corrected = []
    for i, c in enumerate(s[:6]):
        if i in (1, 3, 5):
            if c == 'O': c = '0'
            elif c == 'I': c = '1'
        elif i in (0, 2, 4):
            if c == '0': c = 'O'
        corrected.append(c)
    s = ''.join(corrected)
    return s[:3]

def apply_canonical(street_full, canonical_map):
    m = re.match(r'^(\d+[A-Z]?\s+)(.*)', street_full)
    if m:
        hnum, sname = m.group(1), m.group(2)
    else:
        hnum, sname = '', street_full
    sname = canonical_map.get(sname, sname)
    return hnum + sname


# ── Sidebar: Aliases ──
st.sidebar.header("Street Name Aliases")
st.sidebar.markdown("Add aliases for streets that are the same road but named differently.")

if 'aliases' not in st.session_state:
    st.session_state.aliases = [
        ('PANACHE N SHR RD', 'PANACHE NSHORE RD'),
        ('PANACHE N SHORE RD', 'PANACHE NSHORE RD'),
        ('PANACHE NORTHSHORE RD', 'PANACHE NSHORE RD'),
        ('A PANACHE N SHORE RD', 'PANACHE NSHORE RD'),
        ('A PANACHE N SHR RD', 'PANACHE NSHORE RD'),
        ('PANACHE SHORE RD', 'PANACHE NSHORE RD'),
        ('NORTHSHORE RD', 'PANACHE NSHORE RD'),
        ('N SHORE RD', 'PANACHE NSHORE RD'),
        ('PENACHE NORTHSHORE RD', 'PANACHE NSHORE RD'),
        ('PENACHE N SHORE RD', 'PANACHE NSHORE RD'),
        ('HENNESSY RD', 'HENNESSEY RD'),
        ('OLD SYLVAIN VALLEY HILL RD', 'OLD SLYVAN VALLEY HILL RD'),
        ('LITTLE PENAGE LAKE RD', 'LITTLE PANACHE RD'),
        ('REGIONAL 10 RD', 'REGIONAL RD 10'),
        ('FINDLAY HILL RD', 'FINDLAY RD'),
        ('FINDLAY HILL RD E', 'FINDLAY RD E'),
        ('FINDLAY HILL RD W', 'FINDLAY RD W'),
    ]

with st.sidebar.expander("Current aliases", expanded=False):
    for frm, to in st.session_state.aliases:
        st.text(f"{frm}  →  {to}")

with st.sidebar.form("add_alias"):
    st.markdown("**Add new alias** (use normalized form: uppercase, abbreviated)")
    new_from = st.text_input("From (variant name)")
    new_to = st.text_input("To (canonical name)")
    if st.form_submit_button("Add alias"):
        if new_from and new_to:
            st.session_state.aliases.append((new_from.upper().strip(), new_to.upper().strip()))
            st.success(f"Added: {new_from} → {new_to}")


# ── File uploads ──
st.markdown("---")
upload_col1, upload_col2 = st.columns(2)

with upload_col1:
    st.subheader("1. Hubspot File")
    hub_file = st.file_uploader("Upload Hubspot file (.xlsx / .csv)", type=['xlsx', 'csv'], key='hub')

with upload_col2:
    st.subheader("2. RTA File")
    rta_file = st.file_uploader("Upload RTA file (.xlsx / .csv)", type=['xlsx', 'csv'], key='rta')


def load_file(uploaded_file):
    """Load uploaded file as DataFrame, handling xlsx (with sheet selection) and csv."""
    # SEC-07: Check file size (max 50MB)
    uploaded_file.seek(0, 2)
    size_mb = uploaded_file.tell() / (1024 * 1024)
    uploaded_file.seek(0)
    if size_mb > 50:
        st.error(f"File too large ({size_mb:.1f} MB). Maximum is 50 MB.")
        st.stop()

    if uploaded_file.name.endswith('.csv'):
        return pd.read_csv(uploaded_file), None
    else:
        xl = pd.ExcelFile(uploaded_file)
        return xl, xl.sheet_names


if hub_file and rta_file:

    # ── Load Hubspot ──
    hub_result = load_file(hub_file)
    if isinstance(hub_result[0], pd.ExcelFile):
        hub_xl = hub_result[0]
        hub_sheets = hub_result[1]
        hub_sheet = st.selectbox("Hubspot sheet", hub_sheets, key='hub_sheet') if len(hub_sheets) > 1 else hub_sheets[0]
        df_hub = pd.read_excel(hub_xl, sheet_name=hub_sheet)
    else:
        df_hub = hub_result[0]

    # ── Load RTA ──
    rta_result = load_file(rta_file)
    if isinstance(rta_result[0], pd.ExcelFile):
        rta_xl = rta_result[0]
        rta_sheets = rta_result[1]
        rta_sheet = st.selectbox("RTA sheet", rta_sheets, key='rta_sheet') if len(rta_sheets) > 1 else rta_sheets[0]
        df_rta = pd.read_excel(rta_xl, sheet_name=rta_sheet)
    else:
        df_rta = rta_result[0]

    # SEC-07: Row count guard
    for label, df in [("Hubspot", df_hub), ("RTA", df_rta)]:
        if len(df) > MAX_UPLOAD_ROWS:
            st.error(f"{label} file has {len(df):,} rows. Maximum is {MAX_UPLOAD_ROWS:,}.")
            st.stop()

    st.markdown("---")
    cfg1, cfg2 = st.columns(2)

    # ── Hubspot column config ──
    with cfg1:
        st.subheader("Hubspot Columns")
        st.dataframe(df_hub.head(5), use_container_width=True)
        hub_cols = df_hub.columns.tolist()

        hub_street_col = st.selectbox(
            "Street Address column",
            hub_cols,
            index=hub_cols.index('Street Address') if 'Street Address' in hub_cols else 0,
            key='hub_street'
        )
        hub_pc_col = st.selectbox(
            "Postal Code column",
            hub_cols,
            index=hub_cols.index('Postal Code') if 'Postal Code' in hub_cols else 0,
            key='hub_pc'
        )

    # ── RTA column config ──
    with cfg2:
        st.subheader("RTA Columns")
        st.dataframe(df_rta.head(5), use_container_width=True)
        rta_cols = df_rta.columns.tolist()

        st.markdown("**Address components** (will be combined for matching)")
        rta_addr_no_col = st.selectbox(
            "Address Number column",
            rta_cols,
            index=rta_cols.index('AddressNo') if 'AddressNo' in rta_cols else 0,
            key='rta_addr_no'
        )
        rta_street_col = st.selectbox(
            "Street Name column",
            rta_cols,
            index=rta_cols.index('StreetName') if 'StreetName' in rta_cols else 0,
            key='rta_street'
        )
        rta_locality_col = st.selectbox(
            "Locality / City column",
            rta_cols,
            index=rta_cols.index('Locality') if 'Locality' in rta_cols else 0,
            key='rta_locality'
        )
        rta_pc_col = st.selectbox(
            "Postal Code column",
            rta_cols,
            index=rta_cols.index('PostalCode') if 'PostalCode' in rta_cols else 0,
            key='rta_pc'
        )
        rta_status_col = st.selectbox(
            "RTA Status column (to bring into output)",
            rta_cols,
            index=rta_cols.index('RTA Status') if 'RTA Status' in rta_cols else len(rta_cols)-1,
            key='rta_status'
        )

    # ── Run matching ──
    st.markdown("---")
    enable_no_pc = st.checkbox(
        "Enable risky matching (street-only, ignore postal code mismatch)",
        value=False,
        help="When enabled, addresses that match on street but have different postal codes will be "
             "included in the output (marked red). When disabled, only postal-code-verified matches are exported."
    )
    if st.button("🔍 Run Address Matching", type="primary", use_container_width=True):
        with st.spinner("Matching addresses..."):

            canonical_map = dict(st.session_state.aliases)

            # Build RTA combined full address: "AddressNo StreetName Locality PostalCode"
            df_rta['_rta_full'] = (
                df_rta[rta_addr_no_col].fillna('').astype(str).str.strip() + ' ' +
                df_rta[rta_street_col].fillna('').astype(str).str.strip() + ' ' +
                df_rta[rta_locality_col].fillna('').astype(str).str.strip() + ' ' +
                df_rta[rta_pc_col].fillna('').astype(str).str.strip()
            ).str.strip()

            # Normalize Hubspot
            df_hub['_street'] = df_hub[hub_street_col].fillna('').apply(normalize)
            df_hub['_street_canon'] = df_hub['_street'].apply(lambda s: apply_canonical(s, canonical_map))
            df_hub['_pc'] = df_hub[hub_pc_col].fillna('').apply(norm_pc)

            # Normalize RTA: combine AddressNo + StreetName for matching key
            df_rta['_street'] = (
                df_rta[rta_addr_no_col].fillna('').astype(str) + ' ' +
                df_rta[rta_street_col].fillna('')
            ).apply(normalize)
            df_rta['_street_canon'] = df_rta['_street'].apply(lambda s: apply_canonical(s, canonical_map))
            df_rta['_pc'] = df_rta[rta_pc_col].fillna('').apply(norm_pc)

            # Build key variants
            for df in [df_hub, df_rta]:
                df['_k_exact']     = df['_street']       + '|' + df['_pc']
                df['_k_dir']       = df['_street'].apply(strip_direction) + '|' + df['_pc']
                df['_k_canon']     = df['_street_canon']  + '|' + df['_pc']
                df['_k_canon_dir'] = df['_street_canon'].apply(strip_direction) + '|' + df['_pc']
                # QA-03: Unit-stripped keys (97U2 PIONEER RD -> 97 PIONEER RD)
                df['_k_unit']      = df['_street'].apply(strip_unit) + '|' + df['_pc']
                df['_k_unit_dir']  = df['_street'].apply(strip_unit).apply(strip_direction) + '|' + df['_pc']

            # MED-1: Detect duplicate keys with conflicting statuses
            dup_check = df_rta.groupby('_k_exact')[rta_status_col].nunique()
            conflict_keys = dup_check[dup_check > 1]
            if len(conflict_keys) > 0:
                st.warning(f"**{len(conflict_keys)} RTA address(es) have conflicting statuses.** "
                           f"First match will be used. Review these in the RTA data:")
                conflict_detail = []
                for key in conflict_keys.index[:20]:  # show max 20
                    rows = df_rta[df_rta['_k_exact'] == key][[rta_addr_no_col, rta_street_col, rta_pc_col, rta_status_col]]
                    for _, r in rows.iterrows():
                        conflict_detail.append({
                            'Address': f"{r[rta_addr_no_col]} {r[rta_street_col]}",
                            'PostalCode': r[rta_pc_col],
                            'Status': r[rta_status_col],
                            'Key': key,
                        })
                st.dataframe(pd.DataFrame(conflict_detail), use_container_width=True)

            # Build lookups: key -> (rta_full_address, rta_status) as separate Series
            # For conflicting keys, concatenate ALL addresses and statuses
            conflict_key_set = set(conflict_keys.index)
            conflict_addr_map = {}
            conflict_status_map = {}
            for key in conflict_key_set:
                rows = df_rta[df_rta['_k_exact'] == key]
                addrs = rows['_rta_full'].dropna().unique()
                statuses = rows[rta_status_col].dropna().unique()
                conflict_addr_map[key] = ' | '.join(str(a) for a in addrs)
                conflict_status_map[key] = ' | '.join(str(s) for s in statuses)

            lookup_addr = {}
            lookup_status = {}
            for key_col in ['_k_exact', '_k_dir', '_k_canon', '_k_canon_dir', '_k_unit', '_k_unit_dir']:
                deduped = df_rta.drop_duplicates(subset=key_col).set_index(key_col)
                addr_series = deduped['_rta_full'].copy()
                status_series = deduped[rta_status_col].fillna('').astype(str).copy()
                # Override conflicting keys with all candidates
                for ck in conflict_key_set:
                    if ck in addr_series.index:
                        addr_series[ck] = f"CONFLICT: {conflict_addr_map[ck]}"
                    if ck in status_series.index:
                        status_series[ck] = f"CONFLICT: {conflict_status_map[ck]}"
                lookup_addr[key_col] = addr_series
                lookup_status[key_col] = status_series

            # Initialize output columns
            df_hub['RTA Address'] = pd.Series(dtype='object')
            df_hub['RTA Status'] = pd.Series(dtype='object')
            df_hub['_match_type'] = ''

            passes = [
                ('_k_exact',     'exact'),
                ('_k_dir',       'direction_strip'),
                ('_k_canon',     'fuzzy'),
                ('_k_canon_dir', 'fuzzy'),
                ('_k_unit',      'fuzzy'),
                ('_k_unit_dir',  'fuzzy'),
            ]

            for key_col, mtype in passes:
                unmatched = df_hub['RTA Address'].isna()
                mapped_addr = df_hub.loc[unmatched, key_col].map(lookup_addr[key_col])
                mapped_status = df_hub.loc[unmatched, key_col].map(lookup_status[key_col])
                matched_mask = mapped_addr.notna()
                if matched_mask.any():
                    df_hub.loc[mapped_addr[matched_mask].index, 'RTA Address'] = mapped_addr[matched_mask].values
                    df_hub.loc[mapped_status[matched_mask].index, 'RTA Status'] = mapped_status[matched_mask].values
                    newly_matched = unmatched & df_hub['RTA Address'].notna() & (df_hub['_match_type'] == '')
                    df_hub.loc[newly_matched, '_match_type'] = mtype

            # MED-1: Mark rows that matched a conflicting key with orange
            conflict_key_set = set(conflict_keys.index)
            for idx in df_hub[df_hub['RTA Address'].notna()].index:
                key = df_hub.loc[idx, '_k_exact']
                if key in conflict_key_set and df_hub.loc[idx, '_match_type'] == 'exact':
                    df_hub.loc[idx, '_match_type'] = 'conflict'

            # Pass 5: street-only (no postal code) → RED (opt-in only)
            if not enable_no_pc:
                st.info("Risky matching (street-only, no postal code) is disabled. "
                        "Enable the checkbox above to include these matches.")

            r_lookup_addr = {}
            r_lookup_status_map = {}
            r_lookup_addr_stripped = {}
            r_lookup_status_stripped = {}
            for i in range(len(df_rta)):
                addr_val = df_rta.iloc[i]['_rta_full']
                status_val = str(df_rta.iloc[i].get(rta_status_col, ''))
                for st_key in [df_rta.iloc[i]['_street'], df_rta.iloc[i]['_street_canon']]:
                    if st_key and st_key not in r_lookup_addr:
                        r_lookup_addr[st_key] = addr_val
                        r_lookup_status_map[st_key] = status_val
                for st_key in [strip_direction(df_rta.iloc[i]['_street']), strip_direction(df_rta.iloc[i]['_street_canon'])]:
                    if st_key and st_key not in r_lookup_addr_stripped:
                        r_lookup_addr_stripped[st_key] = addr_val
                        r_lookup_status_stripped[st_key] = status_val

            if enable_no_pc:
                unmatched = df_hub['RTA Address'].isna()
                for idx in df_hub[unmatched].index:
                    h_st = df_hub.loc[idx, '_street']
                    h_st_canon = df_hub.loc[idx, '_street_canon']
                    for lookup_a, lookup_s, key in [
                        (r_lookup_addr, r_lookup_status_map, h_st),
                        (r_lookup_addr, r_lookup_status_map, h_st_canon),
                        (r_lookup_addr_stripped, r_lookup_status_stripped, strip_direction(h_st)),
                        (r_lookup_addr_stripped, r_lookup_status_stripped, strip_direction(h_st_canon)),
                    ]:
                        if key in lookup_a:
                            df_hub.loc[idx, 'RTA Address'] = lookup_a[key]
                            df_hub.loc[idx, 'RTA Status'] = lookup_s.get(key, '')
                            df_hub.loc[idx, '_match_type'] = 'no_pc'
                            break

            # ── Reverse lookup: find RTA addresses NOT in Hubspot ──
            matched_hub_keys = set()
            key_cols_list = ['_k_exact', '_k_dir', '_k_canon', '_k_canon_dir', '_k_unit', '_k_unit_dir']
            for key_col in key_cols_list:
                matched_rows = df_hub[df_hub['RTA Address'].notna()]
                matched_hub_keys.update(matched_rows[key_col].dropna().unique())

            def rta_in_hubspot(row):
                for key_col in key_cols_list:
                    if row[key_col] in matched_hub_keys:
                        return 'Yes'
                return 'No'

            df_rta['In Hubspot'] = df_rta.apply(rta_in_hubspot, axis=1)
            rta_in_hub = (df_rta['In Hubspot'] == 'Yes').sum()
            rta_not_in_hub = (df_rta['In Hubspot'] == 'No').sum()

            # Stats
            exact_count = (df_hub['_match_type'] == 'exact').sum()
            yellow_count = df_hub['_match_type'].isin(['fuzzy', 'direction_strip']).sum()
            orange_count = (df_hub['_match_type'] == 'conflict').sum()
            red_count = (df_hub['_match_type'] == 'no_pc').sum()
            hub_matched = df_hub['RTA Address'].notna().sum()
            hub_unmatched = len(df_hub) - hub_matched

            # ── Dashboard ──
            st.markdown("---")
            st.subheader("📊 Dashboard")

            # Row 1: Side-by-side overview
            d1, d2 = st.columns(2)
            with d1:
                st.markdown("**Hubspot**")
                h1, h2, h3 = st.columns(3)
                h1.metric("Total", len(df_hub))
                h2.metric("Matched", hub_matched)
                h3.metric("Unmatched", hub_unmatched)
            with d2:
                st.markdown("**RTA**")
                r1, r2, r3 = st.columns(3)
                r1.metric("Total", len(df_rta))
                r2.metric("In Hubspot", rta_in_hub)
                r3.metric("Not in Hubspot", rta_not_in_hub)

            # Explain the difference
            if hub_matched != rta_in_hub:
                diff = hub_matched - rta_in_hub
                st.info(f"**Why {hub_matched} vs {rta_in_hub}?** — "
                        f"{diff} Hubspot row(s) map to the same RTA address "
                        f"(duplicate Hubspot entries pointing to one RTA record).")

            # Row 2: Match type breakdown
            st.markdown("**Match breakdown:**")
            b1, b2, b3, b4 = st.columns(4)
            b1.metric("⬜ Exact", exact_count)
            b2.metric("🟨 Fuzzy", yellow_count)
            b3.metric("🟧 Conflict", orange_count)
            b4.metric("🟥 Risky (no PC)", red_count)

            # Show special matches
            special = df_hub[df_hub['_match_type'].isin(['fuzzy', 'direction_strip', 'no_pc', 'conflict'])][
                [hub_street_col, hub_pc_col, 'RTA Address', 'RTA Status', '_match_type']
            ].copy()
            special.columns = ['Street Address', 'Postal Code', 'RTA Address', 'RTA Status', 'Match Type']

            if len(special) > 0:
                st.markdown("**Flagged matches for review:**")

                def highlight_match_type(row):
                    colors = {
                        'no_pc': '#FF6666',
                        'conflict': '#FFA500',
                        'fuzzy': '#FFFF00',
                        'direction_strip': '#FFFF00',
                    }
                    bg = colors.get(row['Match Type'], '#FFFFFF')
                    return [f'background-color: {bg}'] * len(row)

                st.dataframe(special.style.apply(highlight_match_type, axis=1), use_container_width=True)

            # RTA not in Hubspot detail
            if rta_not_in_hub > 0:
                st.markdown(f"**{rta_not_in_hub} RTA addresses not in Hubspot** — "
                            "marked 🟪 purple in the RTA sheet for Redrabbit update.")
                rta_not_matched = df_rta[df_rta['In Hubspot'] == 'No'][
                    [rta_addr_no_col, rta_street_col, rta_locality_col, rta_pc_col, rta_status_col]
                ].head(20)
                st.dataframe(rta_not_matched, use_container_width=True)
                if rta_not_in_hub > 20:
                    st.caption(f"Showing first 20 of {rta_not_in_hub}. Full list in the downloaded Excel.")

            # Preview output
            preview = df_hub[[hub_street_col, hub_pc_col, 'RTA Address', 'RTA Status']].head(20)
            st.markdown("**Hubspot output preview (first 20 rows):**")
            st.dataframe(preview, use_container_width=True)

            # ── Save to Excel with TWO sheets: Hubspot + RTA ──
            match_type = df_hub['_match_type'].copy()
            df_hub_out = df_hub.drop(columns=[c for c in df_hub.columns if c.startswith('_')])
            df_rta_out = df_rta.drop(columns=[c for c in df_rta.columns if c.startswith('_')])

            df_hub_out = sanitize_dataframe(df_hub_out)
            df_rta_out = sanitize_dataframe(df_rta_out)

            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df_hub_out.to_excel(writer, sheet_name='Hubspot', index=False)
                df_rta_out.to_excel(writer, sheet_name='RTA', index=False)
            buffer.seek(0)

            wb = load_workbook(buffer)

            # ── Color Hubspot sheet ──
            ws_hub = wb['Hubspot']
            rta_addr_col_idx = None
            rta_status_col_idx = None
            for cell in ws_hub[1]:
                if cell.value == 'RTA Address':
                    rta_addr_col_idx = cell.column
                elif cell.value == 'RTA Status':
                    rta_status_col_idx = cell.column

            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            red_fill = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')
            orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

            for i, mt in enumerate(match_type):
                if mt in ('fuzzy', 'direction_strip'):
                    fill = yellow_fill
                elif mt == 'no_pc':
                    fill = red_fill
                elif mt == 'conflict':
                    fill = orange_fill
                else:
                    continue
                if rta_addr_col_idx:
                    ws_hub.cell(row=i+2, column=rta_addr_col_idx).fill = fill
                if rta_status_col_idx:
                    ws_hub.cell(row=i+2, column=rta_status_col_idx).fill = fill

            # ── Color RTA sheet: highlight "Not in Hubspot" rows ──
            ws_rta = wb['RTA']
            purple_fill = PatternFill(start_color='D8B4FE', end_color='D8B4FE', fill_type='solid')

            in_hub_col_idx = None
            for cell in ws_rta[1]:
                if cell.value == 'In Hubspot':
                    in_hub_col_idx = cell.column
                    break

            if in_hub_col_idx:
                for row_idx in range(2, ws_rta.max_row + 1):
                    cell = ws_rta.cell(row=row_idx, column=in_hub_col_idx)
                    if cell.value == 'No':
                        for col_idx in range(1, ws_rta.max_column + 1):
                            ws_rta.cell(row=row_idx, column=col_idx).fill = purple_fill

            out_buffer = io.BytesIO()
            wb.save(out_buffer)
            out_buffer.seek(0)

            st.download_button(
                label="📥 Download color-coded Excel (Hubspot + RTA sheets)",
                data=out_buffer,
                file_name="hubspot_rta_matched_output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

            st.markdown("""
            ---
            **Color legend:**

            **Hubspot sheet:**
            - ⬜ **White** — Exact match (street + postal code)
            - 🟨 **Yellow** — Fuzzy match (name alias, direction stripped, spelling variant)
            - 🟧 **Orange** — Exact match but RTA has conflicting statuses for this address
            - 🟥 **Red** — Street matched but postal codes differ — manual verification needed

            **RTA sheet:**
            - 🟪 **Purple** — NOT in Hubspot (needs Redrabbit update)
            - ⬜ **White** — Matched to at least one Hubspot record
            """)

elif hub_file or rta_file:
    st.info("Please upload both files to proceed.")
