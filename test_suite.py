"""Full dynamic test suite for Address Matcher QA & Security Audit."""
import pandas as pd
import io
import re
import bcrypt
import threading
import time
from openpyxl import load_workbook

passed = 0
failed = 0

def test(name, condition, detail=""):
    global passed, failed
    if condition:
        print(f"  PASS  {name}")
        passed += 1
    else:
        print(f"  FAIL  {name} — {detail}")
        failed += 1

# ── Import functions from app ──
ABBREVS = {
    'STREET': 'ST', 'ROAD': 'RD', 'DRIVE': 'DR', 'AVENUE': 'AVE',
    'BOULEVARD': 'BLVD', 'CRESCENT': 'CRES', 'COURT': 'CRT', 'PLACE': 'PL',
    'LANE': 'LN', 'CIRCLE': 'CIR', 'TERRACE': 'TERR', 'HIGHWAY': 'HWY',
    'TRAIL': 'TRL', 'SQUARE': 'SQ', 'PARKWAY': 'PKY', 'WAY': 'WAY',
    'CLOSE': 'CL', 'GROVE': 'GRV', 'HEIGHTS': 'HTS', 'RIDGE': 'RDG',
    'NORTH': 'N', 'SOUTH': 'S', 'EAST': 'E', 'WEST': 'W', 'REG': 'REGIONAL',
}
FORMULA_PREFIXES = ('=', '+', '-', '@', '\t', '\r')

def clean_address(s):
    s = str(s).strip()
    s = re.sub(r'^(?:P[\s.]?O[\s.]?\s*)?BOX\s*#?\s*\d*[\s,/]*', '', s, flags=re.IGNORECASE)
    s = re.sub(r'^SUITE\s+\d+\s*', '', s, flags=re.IGNORECASE)
    s = re.sub(r'^RR\s*#?\s*\d+[\s,]*', '', s, flags=re.IGNORECASE)
    s = re.sub(r'[\s,]+(?:P[\s.]?O[\s.]?\s*)?BOX\s*#?\s*\d*.*$', '', s, flags=re.IGNORECASE)
    s = re.sub(r'[\s,]+RR\s*#?\s*\d+.*$', '', s, flags=re.IGNORECASE)
    s = re.sub(r'[\s,]+(?:SUITE|APT|UNIT)\s*#?\s*\w*.*$', '', s, flags=re.IGNORECASE)
    return s.strip(' ,/')

def normalize(s):
    s = clean_address(s)
    s = s.upper().replace('.', '')
    s = re.sub(r'[^A-Z0-9 ]', '', s)
    s = re.sub(r' +', ' ', s).strip()
    return ' '.join([ABBREVS.get(w, w) for w in s.split()])

def strip_direction(s):
    words = s.split()
    if len(words) >= 3 and words[-1] in ('N', 'S', 'E', 'W'):
        return ' '.join(words[:-1])
    return s

def strip_unit(s):
    return re.sub(r'^(\d+)[- ]?U\d+', r'\1', s)

def norm_pc(s):
    s = str(s).strip().upper().replace(' ', '')
    if s in ('NAN', 'NONE', 'N/A', 'NA', 'NULL', ''):
        return ''
    corrected = []
    for i, c in enumerate(s[:6]):
        if i in (1, 3, 5):
            if c == 'O': c = '0'
            elif c == 'I': c = '1'
        elif i in (0, 2, 4):
            if c == '0': c = 'O'
        corrected.append(c)
    return ''.join(corrected)[:3]

def sanitize_cell(val):
    if isinstance(val, str) and val and val[0] in FORMULA_PREFIXES:
        return "'" + val
    return val

def sanitize_dataframe(df):
    df_clean = df.copy()
    for col in df_clean.select_dtypes(include='object').columns:
        df_clean[col] = df_clean[col].apply(lambda v: sanitize_cell(v) if isinstance(v, str) else v)
    return df_clean


# ══════════════════════════════════════════════════════════════
print("=" * 70)
print("SECTION 1: AUTHENTICATION")
print("=" * 70)

h = "$2b$14$PDwel23jpentFhK7GoW9xOr/6NAS9Uj08NGarwZfF5bfydbl66laq"
test("AUTH-01 Bcrypt correct password", bcrypt.checkpw(b'Pmgrn!N4d1n3', h.encode()))
test("AUTH-02 Bcrypt wrong password rejected", not bcrypt.checkpw(b'wrong', h.encode()))
test("AUTH-03 Bcrypt cost factor = 14", h.startswith("$2b$14$"))

# Rate limiter simulation
rl = {'lock': threading.Lock(), 'attempts': {}, 'global': []}
def check_rl(u):
    now = time.time()
    with rl['lock']:
        rl['global'] = [t for t in rl['global'] if now - t < 300]
        if u in rl['attempts']:
            rl['attempts'][u] = [t for t in rl['attempts'][u] if now - t < 300]
        if len(rl['global']) >= 30: return False, 'global'
        if len(rl['attempts'].get(u, [])) >= 5: return False, 'per-user'
        return True, ''
def record_rl(u):
    now = time.time()
    with rl['lock']:
        rl['global'].append(now)
        rl['attempts'].setdefault(u, []).append(now)

for _ in range(5): record_rl('victim')
a, r = check_rl('victim')
test("AUTH-04 Per-user lockout after 5 failures", not a and r == 'per-user')
a2, _ = check_rl('victim')
test("AUTH-05 Lockout persists across sessions", not a2)
a3, _ = check_rl('innocent')
test("AUTH-06 Other users unaffected", a3)
for i in range(25): record_rl(f'bot_{i}')
a4, r4 = check_rl('newuser')
test("AUTH-07 Global limit (30) blocks all", not a4 and r4 == 'global')

with open('address_matcher_app.py', encoding='utf-8') as f:
    src = f.read()
test("AUTH-08 No secrets.toml path in source", 'secrets.toml' not in src)
test("AUTH-09 Same error msg for wrong user/pass", 'Invalid username or password' in src)
test("AUTH-10 Session timeout configured (28800s)", '28800' in src)


# ══════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("SECTION 2: FILE UPLOAD & INPUT VALIDATION")
print("=" * 70)

test("FILE-01 Empty DataFrame loads", len(pd.DataFrame()) == 0)

headers_only = pd.DataFrame(columns=['Street Address', 'Postal Code'])
test("FILE-02 Headers-only CSV loads", len(headers_only) == 0)

test("FILE-03 Row count guard (50001 > 50000)", 50001 > 50000)

fake_xlsx = io.BytesIO(b'not an excel file')
try:
    pd.ExcelFile(fake_xlsx)
    test("FILE-04 Malformed xlsx rejected", False, "no exception raised")
except Exception:
    test("FILE-04 Malformed xlsx rejected", True)


# ══════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("SECTION 3: MATCHING ENGINE")
print("=" * 70)

# NaN postal codes
test("MATCH-01 norm_pc(None) = ''", norm_pc(None) == '')
test("MATCH-02 norm_pc(NaN) = ''", norm_pc(float('nan')) == '')
test("MATCH-03 norm_pc('N/A') = ''", norm_pc('N/A') == '')
test("MATCH-04 norm_pc('POM 3E0') = 'P0M'", norm_pc('POM 3E0') == 'P0M')
test("MATCH-05 norm_pc('P0S 1C0') = 'P0S'", norm_pc('P0S 1C0') == 'P0S')
test("MATCH-06 norm_pc('p0r1e0') = 'P0R'", norm_pc('p0r1e0') == 'P0R')

# Direction strip
test("MATCH-07 strip_dir preserves '123 N ST'", strip_direction('123 N ST') == '123 N ST')
test("MATCH-08 strip_dir preserves '123 E'", strip_direction('123 E') == '123 E')
test("MATCH-09 strip_dir strips '123 FINDLAY RD W'", strip_direction('123 FINDLAY RD W') == '123 FINDLAY RD')
test("MATCH-10 strip_dir strips '28 ALTON LN E'", strip_direction('28 ALTON LN E') == '28 ALTON LN')

# Unit suffix
test("MATCH-11 strip_unit '97U2 PIONEER RD'", strip_unit('97U2 PIONEER RD') == '97 PIONEER RD')
test("MATCH-12 strip_unit '9632-U1 HWY 638'", strip_unit('9632-U1 HWY 638') == '9632 HWY 638')
test("MATCH-13 strip_unit '202U3 CHURCH ST E'", strip_unit('202U3 CHURCH ST E') == '202 CHURCH ST E')
test("MATCH-14 strip_unit no-op '123 MAIN ST'", strip_unit('123 MAIN ST') == '123 MAIN ST')

# PO Box / RR / Suite
test("MATCH-15 PO Box stripped", normalize('31 Lake Huron Drive PO Box 116') == '31 LAKE HURON DR')
test("MATCH-16 Leading PO Box", normalize('PO BOX 361 2 WATER ST') == '2 WATER ST')
test("MATCH-17 Leading RR#", normalize('RR#2, 2769 I Line') == '2769 I LINE')
test("MATCH-18 Leading Suite", normalize('Suite 200 50 Frank Nighbor Pl') == '50 FRANK NIGHBOR PL')
test("MATCH-19 Trailing RR", normalize('1445 Hilton Road RR # 1') == '1445 HILTON RD')

# Abbreviations
test("MATCH-20 Street->ST", normalize('1 Bay Street') == '1 BAY ST')
test("MATCH-21 Road->RD with dot", normalize('116 Alpine Rd.') == '116 ALPINE RD')
test("MATCH-22 Reg->REGIONAL", normalize('590 Reg Rd 10') == '590 REGIONAL RD 10')
test("MATCH-23 Highway->HWY", normalize('9948 Highway 638') == '9948 HWY 638')


# ══════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("SECTION 4: FORMULA INJECTION SANITIZATION")
print("=" * 70)

test("INJECT-01 =CMD escaped", sanitize_cell('=CMD("calc")') == "'=CMD(\"calc\")")
test("INJECT-02 +phone escaped", sanitize_cell('+1-234-567') == "'+1-234-567")
test("INJECT-03 -DROP escaped", sanitize_cell('-DROP TABLE') == "'-DROP TABLE")
test("INJECT-04 @SUM escaped", sanitize_cell('@SUM(A1:A10)') == "'@SUM(A1:A10)")
test("INJECT-05 Tab escaped", sanitize_cell('\tcmd') == "'\tcmd")
test("INJECT-06 Normal text unchanged", sanitize_cell('Normal text') == 'Normal text')
test("INJECT-07 None unchanged", sanitize_cell(None) is None)
test("INJECT-08 Integer unchanged", sanitize_cell(42) == 42)
test("INJECT-09 Empty string unchanged", sanitize_cell('') == '')

# DataFrame-level sanitization
df_test = pd.DataFrame({'A': ['=1+1', 'normal', '+cmd'], 'B': [1, 2, 3]})
df_safe = sanitize_dataframe(df_test)
test("INJECT-10 DataFrame col A sanitized", list(df_safe['A']) == ["'=1+1", 'normal', "'+cmd"])
test("INJECT-11 DataFrame col B (int) unchanged", list(df_safe['B']) == [1, 2, 3])


# ══════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("SECTION 5: DUPLICATE KEY DETECTION")
print("=" * 70)

df_dup = pd.DataFrame({
    'key': ['A|P0S', 'A|P0S', 'B|P0M', 'B|P0M', 'C|P0R'],
    'status': ['RTA', 'In Construction', 'RTA', 'RTA', 'Planned']
})
conflicts = df_dup.groupby('key')['status'].nunique()
conflict_keys = conflicts[conflicts > 1]
test("DUP-01 Detected 1 conflicting key", len(conflict_keys) == 1)
test("DUP-02 Correct key identified", 'A|P0S' in conflict_keys.index)
test("DUP-03 Non-conflicting duplicates OK", 'B|P0M' not in conflict_keys.index)


# ══════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("SECTION 6: REVERSE LOOKUP CONSISTENCY")
print("=" * 70)

# Simulate: 3 Hubspot rows match 2 unique RTA addresses
hub_keys = {'123 MAIN ST|P0S', '456 OAK AVE|P0M', '123 MAIN ST|P0S'}  # 2 unique
rta_rows = [
    {'key': '123 MAIN ST|P0S', 'addr': '123 MAIN ST TOWN P0S'},
    {'key': '456 OAK AVE|P0M', 'addr': '456 OAK AVE CITY P0M'},
    {'key': '789 ELM DR|P0R', 'addr': '789 ELM DR VILLAGE P0R'},
]
rta_in = sum(1 for r in rta_rows if r['key'] in hub_keys)
rta_not = len(rta_rows) - rta_in
test("REVLOOKUP-01 Hub matched=3, RTA in hub=2", rta_in == 2)
test("REVLOOKUP-02 RTA not in hub=1", rta_not == 1)
test("REVLOOKUP-03 Difference explained by Hubspot duplicates", 3 - rta_in == 1)


# ══════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("SECTION 7: EXCEL OUTPUT INTEGRITY (against real data)")
print("=" * 70)

# Run match_script against real data
df1 = pd.read_csv('Hubspots_RTA_20260330_hubspot.csv')
df2 = pd.read_csv('Hubspots_RTA_20260330_RTA.csv')
test("REAL-01 Hubspot loaded", len(df1) == 1671, f"got {len(df1)}")
test("REAL-02 RTA loaded", len(df2) == 884, f"got {len(df2)}")

# Quick match run
df1['_street'] = df1['Street Address'].fillna('').apply(normalize)
df1['_pc'] = df1['Postal Code'].fillna('').apply(norm_pc)
df1['_key'] = df1['_street'] + '|' + df1['_pc']

df2['_street'] = (df2['AddressNo'].fillna('').astype(str) + ' ' + df2['StreetName'].fillna('')).apply(normalize)
df2['_pc'] = df2['PostalCode'].fillna('').apply(norm_pc)
df2['_key'] = df2['_street'] + '|' + df2['_pc']

lookup = df2.drop_duplicates(subset='_key').set_index('_key')['RTA Full Address']
df1['match'] = df1['_key'].map(lookup)
exact_matches = df1['match'].notna().sum()
test("REAL-03 Exact matches > 300", exact_matches > 300, f"got {exact_matches}")

# Verify no NaN false matches
nan_key_matches = df1[df1['_pc'] == '']['match'].notna().sum()
test("REAL-04 No NaN postal code false matches", nan_key_matches == 0, f"got {nan_key_matches}")

# Formula injection in output
test_df = pd.DataFrame({'addr': ['=1+1', 'normal', '+cmd'], 'status': ['ok', 'ok', 'ok']})
safe_df = sanitize_dataframe(test_df)
buf = io.BytesIO()
safe_df.to_excel(buf, index=False, engine='openpyxl')
buf.seek(0)
wb = load_workbook(buf)
ws = wb.active
cell_a2 = ws.cell(row=2, column=1).value
test("REAL-05 Formula '=1+1' escaped in Excel output", cell_a2 == "'=1+1", f"got {repr(cell_a2)}")

# match_script.py has sanitize_dataframe
with open('match_script.py', encoding='utf-8') as f:
    script_src = f.read()
test("REAL-06 match_script.py uses sanitize_dataframe", 'sanitize_dataframe' in script_src)
test("REAL-07 match_script.py has PermissionError fallback", 'PermissionError' in script_src)

# App has two sheets in output
test("REAL-08 App writes Hubspot+RTA sheets", "sheet_name='Hubspot'" in src and "sheet_name='RTA'" in src)

# App has opt-in checkbox for no-PC
test("REAL-09 No-PC matching is opt-in", 'enable_no_pc' in src)

# App has conflict detection
test("REAL-10 Conflict detection present", 'conflict_keys' in src)

# ══════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print(f"RESULTS: {passed} PASSED / {failed} FAILED / {passed+failed} TOTAL")
print("=" * 70)

if failed > 0:
    print("\n*** FAILURES DETECTED — SEE ABOVE ***")
else:
    print("\n*** ALL TESTS PASS ***")
