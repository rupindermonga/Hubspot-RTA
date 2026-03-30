import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

df1 = pd.read_csv('Hubspots_RTA_20260330_hubspot.csv')
df2 = pd.read_csv('Hubspots_RTA_20260330_RTA.csv')

abbrevs = {
    'STREET': 'ST', 'ROAD': 'RD', 'DRIVE': 'DR', 'AVENUE': 'AVE',
    'BOULEVARD': 'BLVD', 'CRESCENT': 'CRES', 'COURT': 'CRT', 'PLACE': 'PL',
    'LANE': 'LN', 'CIRCLE': 'CIR', 'TERRACE': 'TERR', 'HIGHWAY': 'HWY',
    'TRAIL': 'TRL', 'SQUARE': 'SQ', 'PARKWAY': 'PKY', 'WAY': 'WAY',
    'CLOSE': 'CL', 'GROVE': 'GRV', 'HEIGHTS': 'HTS', 'RIDGE': 'RDG',
    'NORTH': 'N', 'SOUTH': 'S', 'EAST': 'E', 'WEST': 'W',
    'REG': 'REGIONAL',
}

def clean_address(s):
    """Strip PO Box, RR, Suite, Apt, Unit and similar noise from addresses."""
    s = str(s).strip()
    # Handle addresses that START with Box/PO Box (e.g. "Box 125, 14 Water St." or "PO BOX 361 2 WATER ST")
    s = re.sub(r'^(?:P[\s.]?O[\s.]?\s*)?BOX\s*#?\s*\d*[\s,/]*', '', s, flags=re.IGNORECASE)
    # Handle addresses that START with Suite (e.g. "Suite 200 50 Frank Nighbor Pl")
    s = re.sub(r'^SUITE\s+\d+\s*', '', s, flags=re.IGNORECASE)
    # Handle addresses that START with RR (e.g. "RR#2, 2769 I Line" or "RR#2 545")
    s = re.sub(r'^RR\s*#?\s*\d+[\s,]*', '', s, flags=re.IGNORECASE)
    # Strip trailing PO Box / P.O. Box / Box patterns
    s = re.sub(r'[\s,]+(?:P[\s.]?O[\s.]?\s*)?BOX\s*#?\s*\d*.*$', '', s, flags=re.IGNORECASE)
    # Strip trailing RR patterns (RR#1, RR # 1, Rr 2)
    s = re.sub(r'[\s,]+RR\s*#?\s*\d+.*$', '', s, flags=re.IGNORECASE)
    # Strip trailing Suite/Apt/Unit
    s = re.sub(r'[\s,]+(?:SUITE|APT|UNIT)\s*#?\s*\w*.*$', '', s, flags=re.IGNORECASE)
    return s.strip(' ,/')

def normalize(s):
    s = clean_address(s)
    s = s.upper().replace('.', '')
    s = re.sub(r'[^A-Z0-9 ]', '', s)
    s = re.sub(r' +', ' ', s).strip()
    words = s.split()
    words = [abbrevs.get(w, w) for w in words]
    return ' '.join(words)

def strip_direction(s):
    return re.sub(r'\s+[NSEW]$', '', s)

def norm_pc(s):
    s = str(s).strip().upper().replace(' ', '')
    # Fix common typo: letter O vs digit 0 in postal codes
    # Canadian postal codes follow pattern: letter-digit-letter digit-letter-digit
    # Positions 1,3,5 should be digits; 0,2,4 should be letters
    corrected = []
    for i, c in enumerate(s[:6]):
        if i in (1, 3, 5):  # should be digit
            if c == 'O':
                c = '0'
            elif c == 'I':
                c = '1'
        elif i in (0, 2, 4):  # should be letter
            if c == '0':
                c = 'O'
        corrected.append(c)
    s = ''.join(corrected)
    return s[:3]

# Canonical street name mapping (applied to street name portion only)
canonical_map = {
    'PANACHE N SHR RD': 'PANACHE NSHORE RD',
    'PANACHE N SHORE RD': 'PANACHE NSHORE RD',
    'PANACHE NORTHSHORE RD': 'PANACHE NSHORE RD',
    'A PANACHE N SHORE RD': 'PANACHE NSHORE RD',
    'A PANACHE N SHR RD': 'PANACHE NSHORE RD',
    'PANACHE SHORE RD': 'PANACHE NSHORE RD',
    'NORTHSHORE RD': 'PANACHE NSHORE RD',
    'N SHORE RD': 'PANACHE NSHORE RD',
    'PENACHE NORTHSHORE RD': 'PANACHE NSHORE RD',
    'PENACHE N SHORE RD': 'PANACHE NSHORE RD',
    'HENNESSY RD': 'HENNESSEY RD',
    'OLD SYLVAIN VALLEY HILL RD': 'OLD SLYVAN VALLEY HILL RD',
    'LITTLE PENAGE LAKE RD': 'LITTLE PANACHE RD',
    # Regional 10 Road word order variant
    'REGIONAL 10 RD': 'REGIONAL RD 10',
    # Findlay Hill Rd = Findlay Rd (same road in Echo Bay, naming inconsistency)
    'FINDLAY HILL RD': 'FINDLAY RD',
    'FINDLAY HILL RD E': 'FINDLAY RD E',
    'FINDLAY HILL RD W': 'FINDLAY RD W',
}

def apply_canonical(street_full):
    m = re.match(r'^(\d+[A-Z]?\s+)(.*)', street_full)
    if m:
        hnum, sname = m.group(1), m.group(2)
    else:
        hnum, sname = '', street_full
    sname = canonical_map.get(sname, sname)
    return hnum + sname

# --- Hubspot keys ---
df1['_street'] = df1['Street Address'].fillna('').apply(normalize)
df1['_street_canon'] = df1['_street'].apply(apply_canonical)
df1['_pc'] = df1['Postal Code'].fillna('').apply(norm_pc)

# --- RTA keys ---
df2['_street'] = (df2['AddressNo'].fillna('').astype(str) + ' ' + df2['StreetName'].fillna('')).apply(normalize)
df2['_street_canon'] = df2['_street'].apply(apply_canonical)
df2['_pc'] = df2['PostalCode'].fillna('').apply(norm_pc)

# Build all key variants
for df, prefix in [(df1, 'h'), (df2, 'r')]:
    df['_k_exact']       = df['_street']       + '|' + df['_pc']
    df['_k_dir']         = df['_street'].apply(strip_direction) + '|' + df['_pc']
    df['_k_canon']       = df['_street_canon']  + '|' + df['_pc']
    df['_k_canon_dir']   = df['_street_canon'].apply(strip_direction) + '|' + df['_pc']

# Build lookups from RTA
lookups = {}
for key_col in ['_k_exact', '_k_dir', '_k_canon', '_k_canon_dir']:
    lookups[key_col] = df2.drop_duplicates(subset=key_col).set_index(key_col)['RTA Full Address']

# Match in priority order
df1['RTA Address'] = pd.Series(dtype='object')
df1['_match_type'] = ''

passes = [
    ('_k_exact',     'exact'),
    ('_k_dir',       'direction_strip'),
    ('_k_canon',     'fuzzy'),
    ('_k_canon_dir', 'fuzzy'),
]

for key_col, mtype in passes:
    unmatched = df1['RTA Address'].isna()
    mapped = df1.loc[unmatched, key_col].map(lookups[key_col])
    df1.loc[unmatched, 'RTA Address'] = mapped
    newly_matched = unmatched & df1['RTA Address'].notna() & (df1['_match_type'] == '')
    df1.loc[newly_matched, '_match_type'] = mtype
    print(f"After {key_col} ({mtype}): {df1['RTA Address'].notna().sum()} total matches")

fuzzy_count = df1['_match_type'].isin(['fuzzy', 'direction_strip']).sum()
print(f"\nFuzzy/alias matches (to color yellow): {fuzzy_count}")

# --- Pass 5: street-only match (no postal code) for remaining unmatched rows ---
# These are risky matches (different postal codes) -> colored RED
r_lookup_street = {}
r_lookup_street_stripped = {}
for i in range(len(df2)):
    st = df2.iloc[i]['_street']
    st_canon = df2.iloc[i]['_street_canon']
    st_s = strip_direction(st)
    st_canon_s = strip_direction(st_canon)
    rta = df2.iloc[i]['RTA Full Address']
    for key in [st, st_canon]:
        if key and key not in r_lookup_street:
            r_lookup_street[key] = rta
    for key in [st_s, st_canon_s]:
        if key and key not in r_lookup_street_stripped:
            r_lookup_street_stripped[key] = rta

unmatched = df1['RTA Address'].isna()
for idx in df1[unmatched].index:
    h_st = df1.loc[idx, '_street']
    h_st_canon = df1.loc[idx, '_street_canon']
    h_st_s = strip_direction(h_st)
    h_st_canon_s = strip_direction(h_st_canon)
    rta = (r_lookup_street.get(h_st) or r_lookup_street.get(h_st_canon)
           or r_lookup_street_stripped.get(h_st_s) or r_lookup_street_stripped.get(h_st_canon_s))
    if rta:
        df1.loc[idx, 'RTA Address'] = rta
        df1.loc[idx, '_match_type'] = 'no_pc'

no_pc_count = (df1['_match_type'] == 'no_pc').sum()
total = df1['RTA Address'].notna().sum()
print(f"After street-only (no PC): {total} total matches")
print(f"Risky matches (no PC, to color red): {no_pc_count}")

# Show all non-exact matched rows
special_rows = df1[df1['_match_type'].isin(['fuzzy', 'direction_strip', 'no_pc'])][['Street Address', 'Postal Code', 'RTA Address', '_match_type']]
print("\nSpecial matched rows:")
print(special_rows.to_string())

# Save to Excel
match_type = df1['_match_type'].copy()
df1_out = df1.drop(columns=[c for c in df1.columns if c.startswith('_')])
df1_out.to_excel('Hubspots_RTA_20260330_hubspot_updated.xlsx', index=False)

# Color matches: yellow for fuzzy/direction_strip, red for no_pc
wb = load_workbook('Hubspots_RTA_20260330_hubspot_updated.xlsx')
ws = wb.active

rta_col = None
for cell in ws[1]:
    if cell.value == 'RTA Address':
        rta_col = cell.column
        break

yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
red_fill = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')

yellow_count = 0
red_count = 0
for i, mt in enumerate(match_type):
    if mt in ('fuzzy', 'direction_strip'):
        ws.cell(row=i+2, column=rta_col).fill = yellow_fill
        yellow_count += 1
    elif mt == 'no_pc':
        ws.cell(row=i+2, column=rta_col).fill = red_fill
        red_count += 1

wb.save('Hubspots_RTA_20260330_hubspot_updated.xlsx')
print(f"\nColored {yellow_count} cells yellow (fuzzy/alias)")
print(f"Colored {red_count} cells red (street match, postal code mismatch)")
print("Saved: Hubspots_RTA_20260330_hubspot_updated.xlsx")
